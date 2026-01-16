"""
Utility module to read Excel files with formulas and values.
Used to validate calculations against output_check.xlsx
"""
import openpyxl
import pandas as pd
from typing import Dict, List, Optional, Tuple
import numpy as np
from datetime import datetime


def read_excel_with_formulas(file_path: str) -> Tuple[openpyxl.Workbook, openpyxl.Workbook]:
    """
    Read Excel file both with formulas and calculated values.
    Returns: (workbook_with_formulas, workbook_with_values)
    """
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
    return wb_formulas, wb_values


def _make_unique_columns(columns):
    """Make column names unique by appending suffixes to duplicates."""
    seen = {}
    unique_cols = []
    for col in columns:
        col_str = str(col) if col is not None else 'Unnamed'
        if col_str in seen:
            seen[col_str] += 1
            unique_cols.append(f"{col_str}_{seen[col_str]}")
        else:
            seen[col_str] = 0
            unique_cols.append(col_str)
    return unique_cols


def get_sheet_data(wb_formulas: openpyxl.Workbook, wb_values: openpyxl.Workbook, 
                   sheet_name: str, header_row: Optional[int] = None) -> pd.DataFrame:
    """
    Get sheet data as DataFrame, with both formulas and values.
    If header_row is specified, uses that row as header.
    """
    if sheet_name not in wb_formulas.sheetnames:
        return pd.DataFrame()
    
    ws_values = wb_values[sheet_name]
    
    # Convert worksheet to list of lists using openpyxl
    data = []
    for row in ws_values.iter_rows(values_only=True):
        data.append(row)
    
    if not data:
        return pd.DataFrame()
    
    if header_row:
        if len(data) > header_row:
            columns = _make_unique_columns(data[header_row-1])
            df = pd.DataFrame(data[header_row:], columns=columns)
        else:
            df = pd.DataFrame()
    else:
        columns = _make_unique_columns(data[0] if data else [])
        df = pd.DataFrame(data[1:], columns=columns)
    
    return df


def get_cell_formula(wb_formulas: openpyxl.Workbook, sheet_name: str, cell_address: str) -> Optional[str]:
    """Get formula from a specific cell"""
    if sheet_name not in wb_formulas.sheetnames:
        return None
    ws = wb_formulas[sheet_name]
    cell = ws[cell_address]
    return cell.value if cell.data_type == 'f' else None


def get_cell_value(wb_values: openpyxl.Workbook, sheet_name: str, cell_address: str):
    """Get calculated value from a specific cell"""
    if sheet_name not in wb_values.sheetnames:
        return None
    ws = wb_values[sheet_name]
    return ws[cell_address].value


def get_accrual_schedule_expected(wb_values: openpyxl.Workbook) -> pd.DataFrame:
    """
    Extract expected accrual schedule from 'Accured Schedule summ' sheet.
    Returns DataFrame with Month, Domestic, International, GST, Total
    """
    if 'Accured Schedule summ' not in wb_values.sheetnames:
        return pd.DataFrame()
    
    ws = wb_values['Accured Schedule summ']
    
    # Find the row with dates (row 4 has dates)
    date_row = 4
    dates = []
    for col_idx in range(2, 15):  # Columns B to N
        cell = ws.cell(row=date_row, column=col_idx)
        if cell.value and isinstance(cell.value, (pd.Timestamp, datetime)):
            dates.append(cell.value.strftime('%Y-%m'))
        elif cell.value:
            try:
                date_val = pd.to_datetime(cell.value)
                dates.append(date_val.strftime('%Y-%m'))
            except:
                pass
    
    # Get Domestic values (row 5)
    domestic_values = []
    for col_idx in range(2, len(dates) + 2):
        cell = ws.cell(row=5, column=col_idx)
        val = cell.value
        if val is not None:
            # Convert from Cr to actual value (multiply by 10^7)
            domestic_values.append(float(val) * 1e7 if isinstance(val, (int, float)) else 0)
        else:
            domestic_values.append(0)
    
    # Get International values (row 6)
    international_values = []
    for col_idx in range(2, len(dates) + 2):
        cell = ws.cell(row=6, column=col_idx)
        val = cell.value
        if val is not None:
            international_values.append(float(val) * 1e7 if isinstance(val, (int, float)) else 0)
        else:
            international_values.append(0)
    
    # Calculate GST (18% on domestic)
    gst_values = [d * 0.18 for d in domestic_values]
    
    # Create DataFrame
    data = {
        'Month': dates[:len(domestic_values)],
        'Domestic': domestic_values,
        'International': international_values,
        'GST': gst_values,
        'Total': [d + i + g for d, i, g in zip(domestic_values, international_values, gst_values)]
    }
    
    return pd.DataFrame(data)


def get_billing_pattern_expected(wb_values: openpyxl.Workbook) -> Dict:
    """
    Extract expected billing pattern from 'Digital Lic billing pattern' sheet.
    Returns dict with ratios and monthly billing data
    """
    result = {
        'domestic_ratio': None,
        'international_ratio': None,
        'monthly_billing': {}
    }
    
    if 'Digital Lic billing pattern' not in wb_values.sheetnames:
        return result
    
    ws = wb_values['Digital Lic billing pattern']
    
    # Get ratios (row 1 and 2, column B)
    domestic_ratio = ws['B1'].value
    international_ratio = ws['B2'].value
    
    if domestic_ratio is not None:
        result['domestic_ratio'] = float(domestic_ratio)
    if international_ratio is not None:
        result['international_ratio'] = float(international_ratio)
    
    # Get monthly billing data (rows 7-14 have monthly data)
    # Column B has revenue, columns D-O have monthly billing amounts
    for row_idx in range(7, 15):
        month_cell = ws[f'A{row_idx}']
        revenue_cell = ws[f'B{row_idx}']
        
        if month_cell.value and isinstance(month_cell.value, (pd.Timestamp, datetime)):
            month_str = pd.to_datetime(month_cell.value).strftime('%Y-%m')
            revenue = revenue_cell.value if revenue_cell.value else 0
            
            # Get monthly billing amounts (columns D-O, which are months 1-12)
            monthly_amounts = []
            for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
                cell = ws[f'{col_letter}{row_idx}']
                val = cell.value if cell.value is not None else 0
                monthly_amounts.append(float(val) if isinstance(val, (int, float)) else 0)
            
            result['monthly_billing'][month_str] = {
                'revenue': float(revenue) if revenue else 0,
                'monthly_amounts': monthly_amounts
            }
    
    return result


def get_cash_flow_expected(wb_values: openpyxl.Workbook) -> pd.DataFrame:
    """
    Extract expected cash flow from 'CF 24-25' sheet.
    Returns DataFrame with cash flow projection
    """
    if 'CF 24-25' not in wb_values.sheetnames:
        return pd.DataFrame()
    
    # Try to read the sheet - structure may vary
    try:
        ws = wb_values['CF 24-25']
        # Convert worksheet to list of lists using openpyxl
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        if data:
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame()
        # The structure needs to be parsed based on actual layout
        # This is a placeholder - will need to be customized based on actual structure
        return df
    except Exception as e:
        print(f"Error reading CF 24-25 sheet: {e}")
        return pd.DataFrame()


def compare_dataframes(actual: pd.DataFrame, expected: pd.DataFrame, 
                       tolerance: float = 0.01, key_columns: List[str] = None) -> Dict:
    """
    Compare two DataFrames and return comparison results.
    Returns dict with 'match', 'differences', 'summary'
    """
    if actual.empty and expected.empty:
        return {'match': True, 'differences': [], 'summary': 'Both DataFrames are empty'}
    
    if actual.empty or expected.empty:
        return {
            'match': False,
            'differences': ['One DataFrame is empty while the other is not'],
            'summary': 'Shape mismatch'
        }
    
    differences = []
    
    # Compare shapes
    if actual.shape != expected.shape:
        differences.append(f"Shape mismatch: actual {actual.shape} vs expected {expected.shape}")
    
    # Compare key columns if specified
    if key_columns:
        for col in key_columns:
            if col in actual.columns and col in expected.columns:
                actual_vals = actual[col].values
                expected_vals = expected[col].values
                
                for idx, (a, e) in enumerate(zip(actual_vals, expected_vals)):
                    if pd.notna(a) and pd.notna(e):
                        diff = abs(float(a) - float(e))
                        if diff > tolerance:
                            differences.append(f"Row {idx}, Column {col}: actual={a}, expected={e}, diff={diff}")
                    elif pd.notna(a) or pd.notna(e):
                        differences.append(f"Row {idx}, Column {col}: actual={a}, expected={e} (one is NaN)")
    
    return {
        'match': len(differences) == 0,
        'differences': differences,
        'summary': f"{len(differences)} differences found" if differences else "All values match"
    }

